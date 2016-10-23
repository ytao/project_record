import openpyxl
import os
from openpyxl import load_workbook

# 配置信息
project_states=['立项中','填写任务书','找推荐厂家中','签字流程中','等待招标','等待施工','签订合同','施工中','施工完毕','办理验收中','厂家发合同中','验收完毕']
file_name='aaa.xlsx'
sheet_name='Sheet1'
start_row=3
end_row=50


class project_data():
	def __init__(self):
		'''初始化'''
		self.filename=file_name
		if os.path.isfile(file_name)==False:
			self.init_success=False
		else:
			self.wb=load_workbook(self.filename)
			self.ws=self.wb[sheet_name]
			self.start_row=start_row
			self.end_row=end_row

			# 这是所有的信息，读取时候获取全部信息
			self.project_names=[]
			self.project_states=[]

			# 这是单独的每一条数据
			self.record_name=''
			self.record_state=''
			self.record_date='2000-1-1'
			self.init_success=True

	def read_all(self):
		'''获取所有信息'''
		for i in range(self.start_row,self.end_row):
			if str(self.ws.cell(row=i,column=1).value)!='None':
				self.project_names.append(str(self.ws.cell(row=i,column=1).value))
			self.project_states=project_states

		self.record_name=self.ws.cell(row=self.start_row,column=1).value
		self.record_state=self.ws.cell(row=self.start_row,column=2).value
		self.record_date=str(self.ws.cell(row=self.start_row,column=3).value)

	def testWrite(self):
		'''测试功能'''
		self.ws.cell(row=5,column=1).value="END"
		self.wb.save(filename=self.filename)

	def read_record(self,record_name):
		'''填写单行记录'''
		for i in range(self.start_row,self.end_row):
			mstr=str(self.ws.cell(row=i,column=1).value)
			if mstr==record_name:
				self.record_state=str(self.ws.cell(row=i,column=2).value)
				self.record_date=str(self.ws.cell(row=i,column=3).value)
		# return [self.project_state,datetime.datetime.strptime(self.project_rocord_date,'%Y-%m-%d %H:%M:%S')]

	def write_record(self,record_name,record_state,record_date):
		'''填写记录'''
		x=1
		t_found=0
		for i in range(self.start_row,self.end_row):
			if str(self.ws.cell(row=i,column=1).value)==record_name or str(self.ws.cell(row=i,column=1).value)=='None':
				x=i
				t_found=1
				# print(record_name)
				break
		if t_found==0:
			x=end_row+1
		self.ws.cell(row=x,column=1).value=record_name
		self.ws.cell(row=x,column=2).value=record_state
		self.ws.cell(row=x,column=3).value=record_date
		self.wb.save(file_name)

	def delete_record(self,record_name):
		'''删除单行记录'''
		pass
